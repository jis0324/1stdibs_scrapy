# -*- coding: utf-8 -*-
import os
import re
import time
import random
import csv
import json
import scrapy
import traceback
import openpyxl
from scrapy.http import Request
from scrapy import signals

base_dir = os.path.dirname(os.path.abspath(__file__))
out_dir = os.path.dirname(os.path.dirname(base_dir))
output_dir = os.path.join(out_dir, "output")
if not os.path.exists(output_dir):
    os.makedirs(output_dir)

current_time = time.strftime("%Y_%m_%d_%H_%M_%S")
result_file = os.path.join(output_dir, "result_{}.xlsx".format(current_time))
summary_file = os.path.join(output_dir, "@summary.xlsx")
summary_temp_file = os.path.join(output_dir, "@summary_temp.xlsx")

class OstdibsSearchSpider(scrapy.Spider):
    name = '1stdibs_brand_scrapy_crawler'
    allowed_domains = ['1stdibs.com']
    custom_settings = {
        'DOWNLOAD_DELAY': 2,
    }

    def __init__(self, *args, **kwargs):
        # get brand list
        self.brand_list = list()
        self.get_brand_list()

        self.summary_result_file_exist = os.path.exists(summary_file)
        if self.summary_result_file_exist:
            self.format_summary_file()
        else:
            self.create_summary_file()

    @classmethod
    def from_crawler(cls, crawler, *args, **kwargs):
        spider = cls(*args, **kwargs)
        spider._set_crawler(crawler)
        return spider
    
    def _set_crawler(self, crawler):
        self.crawler = crawler
        crawler.signals.connect(self.close, signals.spider_closed)

    # return base_item
    def baseItem(self):
        baseItem = dict()
        baseItem["BRAND"] = ""
        baseItem["PRODUCT_NAME"] = ""
        baseItem["PRODUCT_TYPE"] = ""
        baseItem["QUANTITY"] = ""
        baseItem["DESCRIPTION"] = ""
        baseItem["MATERIALS"] = ""
        baseItem["DIMENSIONS"] = ""
        baseItem["DATE_OF_MANUFACTURE"] = ""
        baseItem["CONDITION"] = ""
        baseItem["CONDITION_NOTES"] = ""
        baseItem["CURRENT_REQUESTED_PRICE"] = ""
        baseItem["SOLD"] = ""
        baseItem["SOLD_PRICE"] = ""
        baseItem["ITEM_LOCATION"] = ""
        baseItem["LIST_DATE"] = ""
        baseItem["SOLD_DATE"] = ""
        baseItem["WEBSITE"] = ""
        baseItem["LISTING_LINK"] = ""
        baseItem["CRAWL_DATE"] = ""
        
        return baseItem

    # get brand list
    def get_brand_list(self):
        brand_txt = base_dir + "/brands.txt"
        with open(brand_txt) as file_object:
            self.brand_list = [row.rstrip('\n') for row in file_object]
    
    # format summary file
    def format_summary_file(self):
        wb_obj = openpyxl.load_workbook(summary_file)

        # from the active attribute 
        sheet_obj = wb_obj.active

        # get max column count
        max_row = sheet_obj.max_row

        for index in range(max_row):
            if index == 0:
                continue
            
            if sheet_obj.cell(row=index, column=13).value == "No":
                sheet_obj.cell(row=index, column=18).value = time.strftime("%d/%m/%Y")

        wb_obj.save(summary_temp_file)

    # create version result file
    def create_version_result_file(self):
        wb_obj = openpyxl.Workbook()
        sheet_obj =  wb_obj.active
        sheet_obj.title = "1stdibs"

        # write fieldnames
        sheet_obj.cell(row=1, column=1).value = "BRAND"
        sheet_obj.cell(row=1, column=2).value = "PRODUCT_NAME"
        sheet_obj.cell(row=1, column=3).value = "PRODUCT_TYPE"
        sheet_obj.cell(row=1, column=4).value = "QUANTITY"
        sheet_obj.cell(row=1, column=5).value = "DESCRIPTION"
        sheet_obj.cell(row=1, column=6).value = "MATERIALS"
        sheet_obj.cell(row=1, column=7).value = "DIMENSIONS"
        sheet_obj.cell(row=1, column=8).value = "DATE_OF_MANUFACTURE"
        sheet_obj.cell(row=1, column=9).value = "CONDITION"
        sheet_obj.cell(row=1, column=10).value = "CONDITION_NOTES"
        sheet_obj.cell(row=1, column=11).value = "CURRENT_REQUESTED_PRICE"
        sheet_obj.cell(row=1, column=12).value = "SOLD"
        sheet_obj.cell(row=1, column=13).value = "SOLD_PRICE"
        sheet_obj.cell(row=1, column=14).value = "ITEM_LOCATION"
        sheet_obj.cell(row=1, column=15).value = "LIST_DATE"
        sheet_obj.cell(row=1, column=16).value = "SOLD_DATE"
        sheet_obj.cell(row=1, column=17).value = "WEBSITE"
        sheet_obj.cell(row=1, column=18).value = "LISTING_LINK"
        sheet_obj.cell(row=1, column=19).value = "CRAWL_DATE"
        wb_obj.save(result_file)

    # update version result file
    def update_version_result_file(self, result_dict):
        wb_obj = openpyxl.load_workbook(result_file)

        # from the active attribute 
        sheet_obj = wb_obj.active

        # get max column count
        max_row = sheet_obj.max_row
        row_index = max_row + 1
        
        sheet_obj.cell(row=row_index, column=1).value = result_dict["BRAND"]
        sheet_obj.cell(row=row_index, column=2).value = result_dict["PRODUCT_NAME"]
        sheet_obj.cell(row=row_index, column=3).value = result_dict["PRODUCT_TYPE"]
        sheet_obj.cell(row=row_index, column=4).value = result_dict["QUANTITY"]
        sheet_obj.cell(row=row_index, column=5).value = result_dict["DESCRIPTION"]
        sheet_obj.cell(row=row_index, column=6).value = result_dict["MATERIALS"]
        sheet_obj.cell(row=row_index, column=7).value = result_dict["DIMENSIONS"]
        sheet_obj.cell(row=row_index, column=8).value = result_dict["DATE_OF_MANUFACTURE"]
        sheet_obj.cell(row=row_index, column=9).value = result_dict["CONDITION"]
        sheet_obj.cell(row=row_index, column=10).value = result_dict["CONDITION_NOTES"]
        sheet_obj.cell(row=row_index, column=11).value = result_dict["CURRENT_REQUESTED_PRICE"]
        sheet_obj.cell(row=row_index, column=12).value = result_dict["SOLD"]
        sheet_obj.cell(row=row_index, column=13).value = result_dict["SOLD_PRICE"]
        sheet_obj.cell(row=row_index, column=14).value = result_dict["ITEM_LOCATION"]
        sheet_obj.cell(row=row_index, column=15).value = result_dict["LIST_DATE"]
        sheet_obj.cell(row=row_index, column=16).value = result_dict["SOLD_DATE"]
        sheet_obj.cell(row=row_index, column=17).value = result_dict["WEBSITE"]
        sheet_obj.cell(row=row_index, column=18).value = result_dict["LISTING_LINK"]
        sheet_obj.cell(row=row_index, column=19).value = result_dict["CRAWL_DATE"]

        wb_obj.save(result_file)

    # create summary file
    def create_summary_file(self):
        wb_obj = openpyxl.Workbook()
        sheet_obj =  wb_obj.active
        sheet_obj.title = "1stdibs"

        # write fieldnames
        sheet_obj.cell(row=1, column=1).value = "BRAND"
        sheet_obj.cell(row=1, column=2).value = "PRODUCT_NAME"
        sheet_obj.cell(row=1, column=3).value = "PRODUCT_TYPE"
        sheet_obj.cell(row=1, column=4).value = "QUANTITY"
        sheet_obj.cell(row=1, column=5).value = "DESCRIPTION"
        sheet_obj.cell(row=1, column=6).value = "MATERIALS"
        sheet_obj.cell(row=1, column=7).value = "DIMENSIONS"
        sheet_obj.cell(row=1, column=8).value = "DATE_OF_MANUFACTURE"
        sheet_obj.cell(row=1, column=9).value = "CONDITION"
        sheet_obj.cell(row=1, column=10).value = "CONDITION_NOTES"
        sheet_obj.cell(row=1, column=11).value = "LAST_CRAWL_REQUESTED_PRICE"
        sheet_obj.cell(row=1, column=12).value = "CURRENT_REQUESTED_PRICE"
        sheet_obj.cell(row=1, column=13).value = "SOLD"
        sheet_obj.cell(row=1, column=14).value = "SOLD_PRICE"
        sheet_obj.cell(row=1, column=15).value = "ITEM_LOCATION"
        sheet_obj.cell(row=1, column=16).value = "LIST_DATE"
        sheet_obj.cell(row=1, column=17).value = "SOLD_DATE"
        sheet_obj.cell(row=1, column=18).value = "OUT_OF_CATALOGUE_DATE"
        sheet_obj.cell(row=1, column=19).value = "WEBSITE"
        sheet_obj.cell(row=1, column=20).value = "LISTING_LINK"
        sheet_obj.cell(row=1, column=21).value = "CRAWL_DATE"

        wb_obj.save(summary_temp_file)

    # update summary file
    def update_summary_file(self, result_dict):
        wb_obj = openpyxl.load_workbook(summary_temp_file)
        sheet_obj = wb_obj.active
        max_row = sheet_obj.max_row

        product_exist = False
        for index in range(max_row):
            if index == 0:
                continue

            if result_dict["LISTING_LINK"] == sheet_obj.cell(row=index, column=19).value:
                product_exist = True
                sheet_obj.cell(row=index, column=11).value = sheet_obj.cell(row=index, column=12).value
                sheet_obj.cell(row=index, column=12).value = result_dict["CURRENT_REQUESTED_PRICE"]
                sheet_obj.cell(row=index, column=18).value = "N/A"

        if not product_exist:
            row_index = max_row + 1

            sheet_obj.cell(row=row_index, column=1).value = result_dict["BRAND"]
            sheet_obj.cell(row=row_index, column=2).value = result_dict["PRODUCT_NAME"]
            sheet_obj.cell(row=row_index, column=3).value = result_dict["PRODUCT_TYPE"]
            sheet_obj.cell(row=row_index, column=4).value = result_dict["QUANTITY"]
            sheet_obj.cell(row=row_index, column=5).value = result_dict["DESCRIPTION"]
            sheet_obj.cell(row=row_index, column=6).value = result_dict["MATERIALS"]
            sheet_obj.cell(row=row_index, column=7).value = result_dict["DIMENSIONS"]
            sheet_obj.cell(row=row_index, column=8).value = result_dict["DATE_OF_MANUFACTURE"]
            sheet_obj.cell(row=row_index, column=9).value = result_dict["CONDITION"]
            sheet_obj.cell(row=row_index, column=10).value = result_dict["CONDITION_NOTES"]
            sheet_obj.cell(row=row_index, column=11).value = ""
            sheet_obj.cell(row=row_index, column=12).value = result_dict["CURRENT_REQUESTED_PRICE"]
            sheet_obj.cell(row=row_index, column=13).value = result_dict["SOLD"]
            sheet_obj.cell(row=row_index, column=14).value = result_dict["SOLD_PRICE"]
            sheet_obj.cell(row=row_index, column=15).value = result_dict["ITEM_LOCATION"]
            sheet_obj.cell(row=row_index, column=16).value = result_dict["LIST_DATE"]
            sheet_obj.cell(row=row_index, column=17).value = result_dict["SOLD_DATE"]
            sheet_obj.cell(row=row_index, column=18).value = "N/A"
            sheet_obj.cell(row=row_index, column=19).value = result_dict["WEBSITE"]
            sheet_obj.cell(row=row_index, column=20).value = result_dict["LISTING_LINK"]
            sheet_obj.cell(row=row_index, column=21).value = result_dict["CRAWL_DATE"]

        wb_obj.save(summary_temp_file)

    # start spider
    def start_requests(self):
        for brand in self.brand_list:
            url_pattern = re.sub(r'\s*\(furniture\)', "", brand.strip())
            meta = {"brand": url_pattern}
            url_pattern = re.sub(r'[^\w]+', "%20", url_pattern.strip())
            item_type = ["21st-pre-owned", "antique-vintage"]
            for item in item_type:
                if "baxter" in brand.lower():
                    url = 'https://www.1stdibs.com/search/furniture/?creator=baxter-furniture&item-type={item_type}&oq={pattern}&q={pattern}&production-time-frame=available-now'.format(item_type=item, pattern=url_pattern.lower())
                else:
                    url = 'https://www.1stdibs.com/search/furniture/?item-type={item_type}&oq={pattern}&q={pattern}&production-time-frame=available-now'.format(item_type=item, pattern=url_pattern.lower())
                print("Requested to ", url)
                yield Request(url, callback=self.parse_listing_pages, meta=meta)

    # parse listing pages
    def parse_listing_pages(self, response):
        brand = response.meta["brand"]
        amount_ele = response.xpath('//*[@id="js-root"]//h1/span/text()').get()

        if amount_ele:
            total_product_count = re.search(r'^\d+', amount_ele.strip()).group()
            page_count = int(int(total_product_count) / 60) + 1
            print("---------------------")
            print("Listing: Found {} products, {} listing pages From {}.".format(total_product_count, page_count, brand))

            for index in range(page_count):
                request_url = "{}&page={}".format(response.url, index+1)
                yield Request(request_url, callback=self.parse_product_urls, meta={"brand": brand})

    # parse product urls
    def parse_product_urls(self, response):
        brand = response.meta["brand"]
        product_urls = response.xpath('//*[@id="js-root"]/div[2]/div[2]/div[2]/div[1]/div//a[@data-tn="item-tile-title-anchor"][@href]/@href').extract()
        print("---------------------")
        print("Products: Found {} products from {}".format(len(product_urls), response.url))
        
        for product_url in product_urls:
            product_url = "https://www.1stdibs.com" + product_url
            yield Request(product_url, callback=self.parse_product_data, meta={"brand": brand})

    # parse product data
    def parse_product_data(self, response):
        try:
            result_dict = self.baseItem()
            
            # brand = response.meta["brand"]
            product_json = ""
            res_content = response.text
            res_content = res_content.split('type="application/ld+json">[')[1]
            res_content = res_content.split("</script>", 1)[0]
            res_content = res_content.rsplit("BreadcrumbList", 1)[0].strip().rsplit("},{" , 1)[0]
            res_content = res_content + "}"
            
            product_json = json.loads(res_content)

            try:
                brand_list = response.xpath("//div[@data-tn='pdp-spec-creator']/span/a//text()").getall()
                brand_find_flag = False
                for product_brand in brand_list:
                    if brand_find_flag:
                        break
                    for identified_brand in self.brand_list:
                        if product_brand.replace(" ", "").lower() in identified_brand.replace(" ", "").lower():
                            result_dict["BRAND"] = identified_brand
                            brand_find_flag = True
                            break
            except:
                pass

            if not result_dict["BRAND"]:
                return

            try:
                result_dict["PRODUCT_NAME"] = product_json["name"]
            except:
                pass

            try:
                result_dict["PRODUCT_TYPE"] = product_json["category"]
            except:
                pass

            try:
                result_dict["QUANTITY"] = 1
            except:
                pass

            try:
                result_dict["DESCRIPTION"] = product_json["description"]
            except:
                pass

            try:
                result_dict["MATERIALS"] = response.xpath('//div[@data-tn="pdp-spec-materials-and-techniques"]//text()').get().strip()
            except:
                pass

            try:
                dimentions_list = response.xpath('//div[@data-tn="pdp-spec-dimensions"]/span//text()').getall()
                result_dict["DIMENSIONS"] = ", ".join(dimentions_list)
            except:
                pass

            try:
                result_dict["DATE_OF_MANUFACTURE"] = response.xpath('//span[@data-tn="pdp-spec-detail-dateOfManufacture"]//text()').get().strip()
            except:
                pass

            try:
                result_dict["CONDITION"] = response.xpath('//span[@data-tn="pdp-spec-detail-condition"]//text()').get().strip()
            except:
                pass
            
            try:
                result_dict["CONDITION_NOTES"] = response.xpath('//span[@data-tn="pdp-spec-detail-conditionDetails"]//text()').get().strip()
            except:
                pass

            try:
                # result_dict["CURRENT_REQUESTED_PRICE"] = product_json["offers"]["price"]
                result_dict["CURRENT_REQUESTED_PRICE"] = response.xpath("//div[@data-tn='price-retail']/span[@data-tn='price-amount']/text()").get().strip()
                if result_dict["CURRENT_REQUESTED_PRICE"] and not re.search(r"\$", result_dict["CURRENT_REQUESTED_PRICE"]):
                    return
            except:
                pass

            try:
                if response.xpath('//div[@data-tn="price-SOLD-price"]/*[name()="svg"]'):
                    result_dict["SOLD"] = "Yes"
                    result_dict["SOLD_PRICE"] = result_dict["CURRENT_REQUESTED_PRICE"]
                    result_dict["CURRENT_REQUESTED_PRICE"] = "N/A"
                    result_dict["SOLD_DATE"] = time.strftime("%d/%m%Y")
                else:
                    result_dict["SOLD"] = "No"
                    result_dict["SOLD_PRICE"] = "N/A"
                    result_dict["SOLD_DATE"] = "N/A"

            except:
                result_dict["SOLD"] = "No"
                result_dict["SOLD_PRICE"] = "N/A"
                result_dict["SOLD_DATE"] = "N/A"
                pass

            try:
                result_dict["ITEM_LOCATION"] = response.xpath('//div[@data-tn="pdp-spec-seller-location"]//text()').get().strip()
            except:
                pass

            try:
                result_dict["LIST_DATE"] = product_json["productionDate"]
            except:
                pass
            
            try:
                result_dict["WEBSITE"] = "1stdibs"
            except:
                pass
            
            result_dict["LISTING_LINK"] = response.url

            try:
                result_dict["CRAWL_DATE"] = time.strftime("%d/%m/%Y")
            except:
                pass

            print("------------------------------")
            print(json.dumps(result_dict, indent=4))

            version_result_file_exist = os.path.exists(result_file)
            if version_result_file_exist:
                self.update_version_result_file(result_dict)
            else:
                self.create_version_result_file()

            self.update_summary_file(result_dict)
                
        except:
            print(traceback.print_exc())
        finally:
            return

    @staticmethod
    def close(spider, reason):
        print()
        print("********************************************")
        print("Do you think the crawler worked successfully?", end="\n\n")
        print("(yes) : Script will merge the summary file and this version result file.")
        print("(no) : Script will not reflect this version result in the summary file.", end="\n\n")
        while True:
            value = input("Type yes/no: ")
            if value == "yes":
                if os.path.exists(summary_temp_file):
                    if os.path.exists(summary_file):
                        os.remove(summary_file)
                    os.rename(summary_temp_file, summary_file)
                break
            elif value == "no":
                if os.path.exists(summary_temp_file):
                    os.remove(summary_temp_file)
                break
            else:
                continue
